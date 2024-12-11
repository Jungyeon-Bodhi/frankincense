#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct  9 15:36:00 2024

@author: Bodhi Global Analysis
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

class WEAI():
    def __init__(self, df, name, tag):
        self.df = df
        self.name = name
        self.tag = tag
        self.score_map = {'input_product_decision':1/10,'autonomy':1/10,'ownership_of_assets':1/15,'purchase_sale_transfer':1/15,
                         'access_credit':1/15,'control_over_income':1/5,'group_membership':1/10,'speaking_public':1/10,
                         'workload':1/10,'leisure':1/10}
        self.columns_indicators = ['input_product_decision','autonomy','ownership_of_assets','purchase_sale_transfer',
                         'access_credit','control_over_income','group_membership','speaking_public','workload','leisure']

    # Calculate individual's empowerment score and label them if they pass the empowerment criteria (equal or more than 0.8)
    def individual_empowerment_score(self):
        df = self.df
        # Individual's Empowerment Score
        df['i_score'] = df.apply(lambda row: round(sum(row[col] * self.score_map[col] for col in self.score_map.keys()), 3), axis=1)
        
        # Label if they pass the empowerment criteria
        df['empowered'] = 0 
        df.loc[df['i_score'] >= 0.8, 'empowered'] = 1
        
        # Calculate the average proportion of domains in which disempowered women experience inadequate achievements
        df2 = df[(df['G1.04']=="female") & (df['empowered']== 0)].copy()
        disempowered_ratios = []
        for col in self.columns_indicators:
            disempowered_ratio = (df2[col] == 0).mean()
            disempowered_ratios.append(disempowered_ratio)
        self.average_disempowered = round(sum(disempowered_ratios) / len(disempowered_ratios),4) # Average proportion of domains in which disempowered women experience inadequate achievements
        self.total_female = len(df[df['G1.04'] == 'female']) # Total number of women
        self.non_empowered_w = len(df[(df['G1.04']=="female") & (df['empowered']==0)]) # Total number not empowered women
        self.empowered_w = len(df[(df['G1.04']=="female") & (df['empowered']==1)]) # Total number empowered women
        self.non_empowered_w_ratio = round(self.non_empowered_w/self.total_female,4) # Percentage of not empowered women
        self.empowered_w_ratio = round(1 - self.non_empowered_w_ratio,4)  # Percentage of empowered women
        self.disempowerment_score = round(self.non_empowered_w_ratio * self.average_disempowered,4) # Value of Disempowerment score
        self.five_de = 1 - self.disempowerment_score # Value of 5DE sub-index
        
        cols = []
        values = []
        for i, col in enumerate(self.columns_indicators):
            cols.append(col)
            values.append(disempowered_ratios[i])
        total_contribution = sum(values)
        if total_contribution > 0:
            for i, col in enumerate(cols):
                values[i] = round(values[i] / total_contribution, 4)
        self.domain_contribution_w = pd.DataFrame({'Disempowerment Countribution Rate (Female)': values}, index=cols)
        
        df3 = df[(df['G1.04']=="male") & (df['empowered']== 0)].copy()
        disempowered_ratios_m = []
        for col in self.columns_indicators:
            disempowered_ratio_m = (df3[col] == 0).mean()
            disempowered_ratios_m.append(disempowered_ratio_m)
        self.average_disempowered_m = round(sum(disempowered_ratios_m) / len(disempowered_ratios_m),4) # Average proportion of domains in which disempowered women experience inadequate achievements
        self.total_male = len(df[df['G1.04'] == 'male']) # Total number of women
        self.non_empowered_m = len(df[(df['G1.04']=="male") & (df['empowered']==0)]) # Total number not empowered men
        self.empowered_m = len(df[(df['G1.04']=="male") & (df['empowered']==1)]) # Total number empowered men
        if self.total_male != 0:
            self.non_empowered_m_ratio = round(self.non_empowered_m/self.total_male,4) # Percentage of not empowered men
        else: self.non_empowered_m_ratio = 0
        self.empowered_m_ratio = round(1 - self.non_empowered_m_ratio,4)  # Percentage of empowered men
        self.disempowerment_score_m = round(self.non_empowered_m_ratio * self.average_disempowered_m,4) # Value of Disempowerment score
        self.five_de_m = 1 - self.disempowerment_score_m # Value of 5DE sub-index
        
        cols = []
        values = []
        for i, col in enumerate(self.columns_indicators):
            cols.append(col)
            values.append(disempowered_ratios_m[i])
        total_contribution = sum(values)
        if total_contribution > 0:
            for i, col in enumerate(cols):
                values[i] = round(values[i] / total_contribution, 4)
        self.domain_contribution_m = pd.DataFrame({'Disempowerment Countribution Rate (Male)': values}, index=cols)
        self.domain_contribution = pd.concat([self.domain_contribution_w, self.domain_contribution_m], axis=1)
        return True
        
    def gender_parity_index(self):
        df = self.df
        df = df[df['G1.06'] == 'dual-adult household']
        df_cleaned = df.dropna(subset=['i_score'])  # Ensure 'i_score' column has no NaN values
        filtered_df = df_cleaned.groupby('G1.01').filter(lambda x: len(x) == 2)
        grouped = filtered_df.groupby('G1.01')  # Group by household identification number
        self.total_respondents = len(df_cleaned) # Total number of dual-adult household respondents
        differences = []  # To store the differences for each group
        
        for name, group in grouped:
            # Separate the scores for males and females
            male_scores = group[group['G1.04'] == 'male'][['G1.01', 'i_score']]
            female_scores = group[group['G1.04'] == 'female'][['G1.01', 'i_score']]
            
            # Ensure there are matching 'G1.01' values between males and females
            merged = pd.merge(male_scores, female_scores, on='G1.01', suffixes=('_male', '_female'))
            
            # Calculate the differences between the scores for matching 'G1.01' values
            if not merged.empty:
                merged['score_diff'] = abs(merged['i_score_male'] - merged['i_score_female'])
                
                # Calculate the average of these differences for the group
                avg_diff = merged['score_diff'].mean()
                differences.append(avg_diff)
        
        # Only calculate the average difference if there are valid differences
        if differences:
            self.average_difference = round(sum(differences) / len(differences),4)  # Average empowerment gap
        else:
            self.average_difference = 0
        
        # Calculate the percent of women not achieving gender parity
        female_higher_count = 0
        total_comparisons = 0
        for name, group in grouped:
            females = group[group['G1.04'] == 'female']
            males = group[group['G1.04'] == 'male']
            if not females.empty and not males.empty:
                for female_score in females['i_score']:
                    for male_score in males['i_score']:
                        total_comparisons += 1
                        if female_score >= male_score:
                            female_higher_count += 1

        if total_comparisons > 0:
            self.gender_parity_ratio = round(female_higher_count / total_comparisons,4)
        else:
            self.gender_parity_ratio = 0
        
        self.gender_not_parity_ratio = round(1 - self.gender_parity_ratio,4)

        # Calculate the GPI Score
        self.gpi = round(1 - (self.gender_not_parity_ratio * self.average_difference),4)
        return True

    def weai_calculation(self):
        self.weai = round((0.9 * self.five_de) + (0.1 * self.gpi),4)
        return True

    def domain_table(self):
        df = self.df[self.df['G1.04'] == 'female']
        list_values = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for idx, col in enumerate(self.columns_indicators):
            count_ones = df[col].sum()
            total_count = df[col].count()
            list_values[idx] = round(count_ones / total_count , 3)
            
        df2 = self.df[self.df['G1.04'] == 'male']
        list_values_m = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for idx, col in enumerate(self.columns_indicators):
            count_ones = df2[col].sum()
            total_count = df2[col].count()
            list_values_m[idx] = round(count_ones / total_count , 3)
            
        domain_table_dic = {'1. Production': {
        ("Input in productive decisions", "Autonomy in production"): [list_values[0], list_values[1]]},
        "2. Resources": {("Ownership of assets", "Purchase, sale, or transfer of assets", "Access to and decisions on credit"): [list_values[2], list_values[3], list_values[4]]},
        "3. Income": {("Control over the use of income",): [list_values[5]]},"4. Leadership": {("Group membership", "Speaking in public"): [list_values[6], list_values[7]]},
        "5. Time allocation": {("Workload", "Leisure"): [list_values[8], list_values[9]]}}
        
        domain_table_dic_m = {'1. Production': {
        ("Input in productive decisions", "Autonomy in production"): [list_values_m[0], list_values_m[1]]},
        "2. Resources": {("Ownership of assets", "Purchase, sale, or transfer of assets", "Access to and decisions on credit"): [list_values_m[2], list_values_m[3], list_values_m[4]]},
        "3. Income": {("Control over the use of income",): [list_values_m[5]]},"4. Leadership": {("Group membership", "Speaking in public"): [list_values_m[6], list_values_m[7]]},
        "5. Time allocation": {("Workload", "Leisure"): [list_values_m[8], list_values_m[9]]}}
        
        
        rows = []
        
        for domain, indicators in domain_table_dic.items():
            for indicator_list, values in indicators.items():
                for indicator, value in zip(indicator_list, values):
                    rows.append([domain, indicator, value])
                    
        rows2 = []
        
        for domain, indicators in domain_table_dic_m.items():
            for indicator_list, values in indicators.items():
                for indicator, value in zip(indicator_list, values):
                    rows2.append([domain, indicator, value])

        df = pd.DataFrame(rows, columns=['Domain', 'Indicator', 'Adequacy (Female)'])
        df['Inadequacy (Female)'] = 1 - df['Adequacy (Female)']
        df2 = pd.DataFrame(rows2, columns=['Domain', 'Indicator', 'Adequacy (Male)'])
        df = pd.merge(df, df2, on=['Domain', 'Indicator'])
        df['Inadequacy (Male)'] = 1 - df['Adequacy (Male)']

        self.domain_table = df
        return self.domain_table
        
    def weai_table(self):
        indicator_list = ['5DE','Disempowerment Score','N', "Average % of disempowered women's inadequate achievements","% of women achieving empowerment","% of women not achieving empowerment",
                         "GPI Score", "N (Number of dual-adult households)", "% of women achieving gender parity", 
                          "% of women not achieving gender parity", "Average empowerment gap", "WEAI Score"]
        detail_list = ["The 5DE sub-index assesses the extent of women's empowerment in the five domains. A higher number reflects greater empowerment",
        '-','Total number of women interviewed',"Average proportion of domains in which disempowered women experience inadequate achievements",'Percentage of women with 5DE scores of 80% or more',
        'Percentage of women with 5DE scores of less than 80%',
        'The GPI sub-index measures the inequality in 5DE scores between the primary adult male decisionmakers and primary adult female decisionmakers in the households',
        'The number of households with both a primary male and primary female decisionmaker',
        'Percentage of women who have 5DE scores equal to or higher than those of the primary adult males in their households',
        'Percentage of women who have 5DE scores lower than those of the primary adult males in their households',
        'For women lacking parity, the average percentage shortfall they experience relative to the males in their household',
        "Women's Empowerment in Agriculture Index"]
        value_list_f = [self.five_de, self.disempowerment_score, self.total_female, self.average_disempowered, self.empowered_w_ratio,
                     self.non_empowered_w_ratio, self.gpi, self.total_respondents, self.gender_parity_ratio, self.gender_not_parity_ratio,
                     self.average_difference, self.weai]
        value_list_m = [self.five_de_m, self.disempowerment_score_m, self.total_male, self.average_disempowered_m, self.empowered_m_ratio,
                     self.non_empowered_m_ratio, '-', '-', '-', '-','-', '-']
        indicators_to_table = {"Indicator": indicator_list, "Detail": detail_list,"Female": value_list_f, "Male": value_list_m}
        table = pd.DataFrame(indicators_to_table)
        self.weai_table = table
        return self.weai_table

    def save_tables(self, file_path):
        empty_df1 = pd.DataFrame()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            empty_df1.to_excel(writer, sheet_name='Delete this sheet', index=False)
            
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            self.weai_table.to_excel(writer, sheet_name=self.name, index=True, header=True)
            startrow = self.weai_table.shape[0] + 2
            self.domain_table.to_excel(writer, sheet_name=self.name, startrow=startrow, index=True, header=True)
            startrow = self.domain_table.shape[0] + startrow + 2
            self.domain_contribution.to_excel(writer, sheet_name=self.name, startrow=startrow, index=True, header=True)
                
        wb = load_workbook(file_path)
        ws = wb[self.name]
        ws.insert_rows(1)
        ws['B1'] = self.name
        ws['B1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(file_path)
        
    def analysis(self, file_path):
        self.individual_empowerment_score()
        self.gender_parity_index()
        self.weai_calculation()
        self.weai_table()
        self.domain_table()
        self.save_tables(file_path)
        self.df.to_csv(f"data/{self.tag} final.csv", index=False)
        self.df.to_excel(f"data/{self.tag} final.xlsx", index=False)
        
    def empowerment_breakdown(self, group_by_dict):
        # Get the column and group name from the dictionary
        group_by_column = list(group_by_dict.keys())[0]
        group_name = group_by_dict[group_by_column]
        
        # Get the unique values in the specified group column
        unique_values = self.df[group_by_column].unique()
        unique_values = sorted(unique_values)
        
        # Initialize an empty list to store rows of metrics for each group
        metrics_list = []
        
        # Store the original DataFrame
        df_original = self.df
        
        # For each unique value in the specified column
        for value in unique_values:
            # Reset the DataFrame
            self.df = df_original
            
            # Filter the DataFrame by the specific value in the group_by_column
            filtered_df = self.df[self.df[group_by_column] == value].copy()
            
            # Temporarily assign the filtered DataFrame
            self.df = filtered_df
            
            # Calculate individual empowerment score
            self.individual_empowerment_score()
        
            # Calculate gender parity index
            self.gender_parity_index()
        
            # Calculate WEAI (Women's Empowerment in Agriculture Index)
            self.weai_calculation()
            
            # Create the dictionary for the row with the group name and calculated metrics
            row = {
                f"{group_name}": value,
                "5DE (Female)": self.five_de,
                "5DE (Male)": self.five_de_m,
                "Disempowerment Score (Female)": self.disempowerment_score,
                "Disempowerment Score (Male)": self.disempowerment_score_m,
                "N (Female)": self.total_female,
                "N (Male)": self.total_male,
                "Average % of disempowered women's inadequate achievements":self.average_disempowered,
                "Average % of disempowered men's inadequate achievements":self.average_disempowered_m,
                "% of women achieving empowerment": self.empowered_w_ratio,
                "% of men achieving empowerment": self.empowered_m_ratio,
                "% of women not achieving empowerment": self.non_empowered_w_ratio,
                "% of men not achieving empowerment": self.non_empowered_m_ratio,
                "GPI Score": self.gpi,
                "N (Number of dual-adult households)": self.total_respondents,
                "% of women achieving gender parity": self.gender_parity_ratio,
                "% of women not achieving gender parity": self.gender_not_parity_ratio,
                "Average empowerment gap": self.average_difference,
                "WEAI Score": self.weai
            }
            
            # Append the row to the metrics_list
            metrics_list.append(row)
        
        # Convert the list of rows into a DataFrame
        results_df = pd.DataFrame(metrics_list)
        
        # Restore the original DataFrame
        self.df = df_original
        
        # Return the final DataFrame with the breakdown of metrics
        return results_df
    
    def breakdown_save(self, breakdown, file_path):
        # Iterate through the breakdown dictionary
        for idx, group_by_dict in breakdown.items():
            # Call the empowerment_breakdown function with the group_by_dict
            result_df = self.empowerment_breakdown(group_by_dict)
            
            # Dynamically assign the result DataFrame to an attribute
            # The attribute name will be result_df1, result_df2, result_df3, etc.
            result_df_name = f"result_df{idx}"
            setattr(self, result_df_name, result_df)

        empty_df1 = pd.DataFrame()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            empty_df1.to_excel(writer, sheet_name='Delete this sheet', index=False)
            
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Start writing the DataFrames with a row gap
            row_start = 0
            for idx in range(1, len(breakdown) + 1):
                result_df = getattr(self, f"result_df{idx}")  # Access the dynamically created DataFrame
                result_df.to_excel(writer, sheet_name='Results', startrow=row_start, index=False)
                row_start += len(result_df) + 2  # Add the number of rows in the current DataFrame plus 2 for spacing

        wb = load_workbook(file_path)
        ws = wb['Results']
        ws.insert_rows(1)
        ws['A1'] = "Breakdown WEAI"
        ws['A1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(file_path)